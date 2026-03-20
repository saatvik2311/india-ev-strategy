"""
reports/generate_report.py
Generates a professional Excel report with 4 sheets + embedded charts.
"""
import os
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

REPORT_DIR  = Path(__file__).parent.parent / "outputs"
CHART_DIR   = REPORT_DIR / "charts"
REPORT_PATH = REPORT_DIR / "India_EV_Investment_Report.xlsx"

# ── Style helpers ─────────────────────────────────────────────────────────────
def _hdr(ws, row, col, value, bg="1B4F8A", fg="FFFFFF", bold=True, size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill    = PatternFill("solid", fgColor=bg)
    cell.font    = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell

def _val(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell

def _thick_border():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

TIER_COLORS = {
    "Tier 1 (Premium)":  "1B4F8A",
    "Tier 2 (High)":     "27AE60",
    "Tier 3 (Moderate)": "E67E22",
    "Tier 4 (Low)":      "C0392B",
}


def _write_sheet_exec_summary(wb, scores_df, base_fin, mc_df):
    ws = wb.create_sheet("Executive Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55

    _hdr(ws, 1, 1, "INDIA EV + RENEWABLE INFRASTRUCTURE INVESTMENT STRATEGY",
         bg="1B4F8A", size=13)
    ws.merge_cells("A1:B1")
    _hdr(ws, 2, 1, "Executive Summary — March 2026", bg="2C3E50", size=10)
    ws.merge_cells("A2:B2")

    top3 = scores_df.head(3)
    insights = [
        ("Context",       "India is targeting 30% EV penetration by 2030, creating ₹4.5+ lakh crore infrastructure investment opportunity."),
        ("Methodology",   "Analysis uses 5 quantitative models: demand forecasting, financial modelling (Monte Carlo), MCDA scoring, grid stress, and policy incentives."),
        ("Top Recommendation", f"Prioritize Phase 1 in {top3.iloc[0]['state']} (Score: {top3.iloc[0]['composite_score']:.0f}/100), "
                               f"{top3.iloc[1]['state']} (Score: {top3.iloc[1]['composite_score']:.0f}/100), "
                               f"and {top3.iloc[2]['state']} (Score: {top3.iloc[2]['composite_score']:.0f}/100)."),
        ("Financial View", f"Best base IRR: {base_fin.sort_values('irr_pct', ascending=False).iloc[0]['state']} at "
                           f"{base_fin['irr_pct'].max():.1f}%. Monte Carlo P50 median: "
                           f"{mc_df['irr_p50'].max():.1f}% (P10–P90 range across top states)."),
        ("Risk View",     "Key risks: tariff revision (±15% sensitivity), EV adoption slowdown (Low scenario reduces IRR by ~3–5%), grid capacity constraints in high-stress states."),
        ("Phase 1 Action", f"Allocate capital to Tier 1 states ({', '.join(scores_df[scores_df['tier']=='Tier 1 (Premium)']['state'].tolist()[:3])}). "
                           f"Target 10–50 MW installations per state in Year 1."),
    ]
    for r_offset, (label, text) in enumerate(insights, start=4):
        _hdr(ws, r_offset, 1, label, bg="34495E", size=9)
        c = ws.cell(row=r_offset, column=2, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_offset].height = 40

    # Embed composite score chart if available
    chart_path = CHART_DIR / "1_composite_scores.png"
    if chart_path.exists():
        img = XLImage(str(chart_path))
        img.width, img.height = 560, 370
        ws.add_image(img, "A11")


def _write_sheet_rankings(wb, scores_df):
    ws = wb.create_sheet("State Rankings")
    headers = ["Rank", "State", "Score (/100)", "Tier", "IRR (%)",
               "Demand 2029 (MWh)", "Solar GHI", "Policy Score", "Urban %"]
    col_w   = [6, 18, 13, 18, 10, 18, 12, 14, 12]
    for i, (h, w) in enumerate(zip(headers, col_w), 1):
        _hdr(ws, 1, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w

    for r_idx, (_, row) in enumerate(scores_df.iterrows(), start=2):
        tier_color = TIER_COLORS.get(str(row["tier"]), "607D8B")
        vals = [r_idx - 1, row["state"], row["composite_score"], str(row["tier"]),
                row["irr_pct"], int(row["demand_mwh_2029"]),
                row["solar_ghi_kwh_m2_day"], row["ev_policy_score"], row["urban_pct"]]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx == 4:
                cell.fill = PatternFill("solid", fgColor=tier_color)
                cell.font = Font(color="FFFFFF", bold=True, size=9)

    # Conditional formatting on Score column
    score_col = f"C2:C{len(scores_df)+1}"
    ws.conditional_formatting.add(score_col, ColorScaleRule(
        start_type="min", start_color="C0392B",
        mid_type="percentile", mid_value=50, mid_color="F39C12",
        end_type="max", end_color="27AE60"
    ))


def _write_sheet_financials(wb, base_fin, policy_fin, mc_df):
    ws = wb.create_sheet("Financial Model")
    headers = ["State", "Solar Yield\n(MWh/MW/yr)", "CAPEX\n(₹ Cr)", "Base IRR\n(%)",
               "Base NPV\n(₹ Cr)", "Policy IRR\n(%)", "Policy NPV\n(₹ Cr)",
               "Payback\n(Yrs)", "MC P10\n(%)", "MC P50\n(%)", "MC P90\n(%)"]
    col_w = [18, 14, 12, 12, 13, 12, 14, 12, 11, 11, 11]
    for i, (h, w) in enumerate(zip(headers, col_w), 1):
        _hdr(ws, 1, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w

    merged = base_fin.merge(policy_fin[["state","irr_pct_policy","npv_cr_policy"]], on="state")
    merged = merged.merge(mc_df[["state","irr_p10","irr_p50","irr_p90"]], on="state")

    for r_idx, (_, row) in enumerate(merged.iterrows(), start=2):
        vals = [row["state"], row["solar_yield_mwh_per_mw"], row["capex_cr"],
                row["irr_pct"], row["npv_cr"], row["irr_pct_policy"], row["npv_cr_policy"],
                row["payback_yrs"], row["irr_p10"], row["irr_p50"], row["irr_p90"]]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data bar on IRR column
    irr_range = f"D2:D{len(merged)+1}"
    ws.conditional_formatting.add(irr_range, DataBarRule(
        start_type="min", end_type="max",
        color="1B4F8A", showValue=True
    ))

    # Embed Monte Carlo histogram
    mc_chart = CHART_DIR / "7_monte_carlo_histogram.png"
    if mc_chart.exists():
        img = XLImage(str(mc_chart))
        img.width, img.height = 600, 240
        ws.add_image(img, f"A{len(merged)+4}")


def _write_sheet_scenarios(wb, scenarios_data: dict):
    """scenarios_data: {'Low': df, 'Base': df, 'High': df} each with state, irr_pct, demand_mwh_2029"""
    ws = wb.create_sheet("Scenario Analysis")
    _hdr(ws, 1, 1, "Scenario Comparison — IRR (%) by State", bg="1B4F8A", size=11)
    ws.merge_cells("A1:D1")

    _hdr(ws, 2, 1, "State")
    _hdr(ws, 2, 2, "Low Scenario IRR (%)", bg="C0392B")
    _hdr(ws, 2, 3, "Base Scenario IRR (%)", bg="1B4F8A")
    _hdr(ws, 2, 4, "High Scenario IRR (%)", bg="27AE60")
    for col in "ABCD":
        ws.column_dimensions[col].width = 22

    states = scenarios_data["Base"]["state"].tolist()
    for r_idx, state in enumerate(states, start=3):
        ws.cell(row=r_idx, column=1, value=state).alignment = Alignment(horizontal="left")
        for c_idx, scenario in enumerate(["Low", "Base", "High"], start=2):
            sdf = scenarios_data[scenario]
            row = sdf[sdf["state"] == state]
            val = float(row["irr_pct"].values[0]) if not row.empty else None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="center")


def generate_excel_report(scores_df, base_fin, policy_fin, mc_df,
                           scenarios_data: dict, in_memory=False):
    print("\n── Generating Excel report ─────────────────────────────────────────")
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    _write_sheet_exec_summary(wb, scores_df, base_fin, mc_df)
    _write_sheet_rankings(wb, scores_df)
    _write_sheet_financials(wb, base_fin, policy_fin, mc_df)
    _write_sheet_scenarios(wb, scenarios_data)

    if in_memory:
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    else:
        wb.save(REPORT_PATH)
        print(f"  ✓  India_EV_Investment_Report.xlsx → {REPORT_PATH}")
        return REPORT_PATH
